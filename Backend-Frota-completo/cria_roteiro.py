import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image
from pathlib import Path
from datetime import datetime
import schemas
import os
from dotenv import load_dotenv

load_dotenv()

UPLOAD_DIR = os.getenv("ROTEIROS_UPLOAD_DIR") 

def Roteiro(roteiro: schemas.Roteiro):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diário de Bordo"

    ws.sheet_view.showGridLines = False

    larguras = {
        "A": 10, "B": 23, "C": 12, "D": 15, "E": 40,
        "F": 13, "G": 13, "H": 13, "I": 13, "J": 15
    }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


    for i in range(1, 9):
        ws.row_dimensions[i].height = 23


    ws.row_dimensions[9].height = 16
    ws.row_dimensions[10].height = 16

    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    preenchimento_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=1, max_row=26, min_col=1, max_col=10):
        for cell in row:
            cell.border = borda
            cell.alignment = Alignment(vertical="center", horizontal="center")


    ws.merge_cells("A1:B2")
    ws.merge_cells("C1:H2")   
    ws.merge_cells("I1:J2")   
    ws.merge_cells("A3:J3")
    ws.merge_cells("A4:C4")
    ws.merge_cells("D4:E4")
    ws.merge_cells("F4:G4")
    ws.merge_cells("H4:J4")
    ws.merge_cells("A5:E5")
    ws.merge_cells("F5:J5")
    ws.merge_cells("A6:E6")
    ws.merge_cells("F6:J6")
    ws.merge_cells("A7:E7")
    ws.merge_cells("F7:J7")
    ws.merge_cells("A8:J8") 
    ws.merge_cells("A9:A10")  
    ws.merge_cells("B9:B10")  
    ws.merge_cells("C9:E10")  
    ws.merge_cells("F9:F10")  
    ws.merge_cells("G9:G10")  
    ws.merge_cells("H9:J10")  

    for r in range(11, 21):
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)  # Endereço
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10) # Kilometragem

    ws.merge_cells("C21:E21")
    ws.merge_cells("H21:J21")
    ws.merge_cells("C22:E22")
    ws.merge_cells("H22:J22")
    ws.merge_cells("C23:E23")
    ws.merge_cells("H23:J23")
    ws.merge_cells("B24:G24")
    ws.merge_cells("B25:G25")
    ws.merge_cells("H24:J24")
    ws.merge_cells("H25:J26")
    ws.merge_cells("A24:A26")
    ws.merge_cells("B26:G26")

    # Pegando a data atual formatada (Exemplo: 08-07-2026)
    # data_atual = datetime.now().strftime("%d-%m-%Y")

    date_obj = roteiro.createdat.strftime("%d/%m/%Y")

    print(date_obj)
    print(roteiro.plate)

    ws["C1"] = "DIÁRIO DE BORDO"
    ws["I1"] = "DATA:"
    ws["A4"] = f"MOTORISTA: {roteiro.driver_name}"
    ws["D4"] = f"AJUDANTE: {roteiro.helper}"
    ws["F4"] = "HORA SAÍDA:"
    ws["H4"] = "HORA CHEGADA:"
    ws["A5"] = f"PLACA: {roteiro.plate}"
    ws["F5"] = "KM INICIAL:"

    # Deixei a data da célula A6 dinâmica também para combinar com o arquivo!
    ws["A6"] = f"DATA: {date_obj}" 

    ws["F6"] = "KM FINAL:"
    ws["F7"] = "HORA DE ALMOÇO:"
    ws["A9"] = "NF ou\nPedido"
    ws["B9"] = "RAZAO SOCIAL"
    ws["C9"] = "ENDEREÇO"
    ws["F9"] = "HORA\nCHEGADA"
    ws["G9"] = "HORA\nSAÍDA"
    ws["H9"] = "KILOMETRAGEM"     


    print(roteiro.routes_items)
    for v in roteiro.routes_items:
        ws[f"A{10+v.sequence}"] = f"{v.ordernumber if v.ordernumber is not None else v.reason}"
        ws[f"B{10+v.sequence}"] = f"{v.clientname.strip() if v.clientname is not None else "..."}"
        ws[f"C{10+v.sequence}"] = f"{v.address}, {v.address_number} - {v.neighborhood}, {v.city} - {v.state}, {v.zipcode}"



    # ws["A12"] = "ALARME"
    # ws["B12"] = "VIPEER"
    # ws["C12"] = "Av. Guapira, 773 - Tucuruvi, São Paulo - SP, 02265-001"

    # ws["A13"] = "42.753"
    # ws["B13"] = "INFINITO ( SAC )"
    # ws["C13"] = "Tv. Japauim, 33 - Vila Nova Galvão, São Paulo - SP, 02280-290"

    # ws["A14"] = "RETIRA"
    # ws["B14"] = "DESLI"
    # ws["C14"] = "R. Acopiara, 148 - Vila Nova Cumbica, Guarulhos - SP, 07230-050"

    # ws["A15"] = "10.389"
    # ws["B15"] = "TRUCKVAN"
    # ws["C15"] = "Estr. Velha Guarulhos-Arujá, 950 - Jardim Alamo, Guarulhos - SP, 07250-155"

    # ws["A16"] = "10.389"
    # ws["B16"] = "TRUCKVAN"
    # ws["C16"] = "Estr. Velha Guarulhos-Arujá, 950 - Jardim Alamo, Guarulhos - SP, 07250-155"


    ws["A24"] = "OBS:"
    ws["H24"] = "MOTORISTA:"
    ws["H25"] = "AJUDANTE:"

    for row in ws["C1:H2"]:
        for cell in row:
            cell.font = Font(bold=True, size=18)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    intervalos_vermelhos = ["A4:C4", "D4:E4", "F4:G4", "H4:J4", "A5:E5", "A6:E6"]
    for intervalo in intervalos_vermelhos:
        for row in ws[intervalo]:
            for cell in row:
                cell.font = Font(color="FF0000", bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in ws["A9:J10"]:
        for cell in row:
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(11, 21):
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)

            if col in (1, 2):  
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True                   
                )

    for r in range(24, 27):
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            
            if col <= 7 or r == 26:
                cell.fill = preenchimento_amarelo
            
                if 2 <= col <= 7:
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                if col == 1:
                    cell.font = Font(bold=True, size=10)
            else:

                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True,size=10)                  

    logo = Path(__file__).parent / "logo.png"
    if logo.exists():
        img = Image(str(logo))
        img.width = 256  
        img.height = 88
        ws.add_image(img, "A1")

    nome_arquivo = f"/var/www/uploads/roteiros/planilhas/diario_bordo_{roteiro.id}.xlsx"
    arquivo = Path(__file__).parent / nome_arquivo


    # contador = 1
    # while arquivo.exists():
    #     nome_arquivo = f"/var/www/uploads/roteiros/diario_bordo_{roteiro.id}.xlsx"
    #     arquivo = Path(__file__).parent / nome_arquivo
    #     contador += 1

    wb.save(arquivo)

    return nome_arquivo

